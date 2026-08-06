"""
GL Reconciliation: Export utilities

Exports reconciliation results to Excel format for accounting review.
"""

from typing import List, Optional, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reconciler import MatchResult


def export_results_to_excel(
    buffer,
    results: List[MatchResult],
    filename: str = "GL_Reconciliation_Report.xlsx"
) -> None:
    """
    Export reconciliation results to Excel workbook.

    Creates a report with columns:
    Account Code | Vendor | Invoice # | Date | Amount | Action

    Rows are sorted by GL code, then by vendor, then by invoice number.
    Only includes unmatched invoices (missing RVW, missing Craftable, or amount mismatch).

    Action is a starting suggestion for accounting, not a final decision:
    - "accrue": invoice exists in Craftable but not yet posted to RVW
    - "missing": invoice exists in RVW but Craftable has no record of it
    - "reclass": same invoice/vendor, but Craftable and RVW allocated it to
      different GL accounts (a likely miscategorization)
    - blank: amount differs on the same GL code in both systems - accounting
      needs to determine the correct figure, hence no auto-suggested action

    Args:
        buffer: BytesIO buffer to write Excel file to
        results: List of MatchResult objects
        filename: Filename (for reference, not written to buffer)
    """
    
    # Filter to only unmatched results
    unmatched = [r for r in results if r.match_type != 'matched']
    
    # Sort results for display: by GL code, then by date (per BK's spec)
    unmatched.sort(key=lambda r: (r.gl_code, r.date, r.invoice_number))
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Discrepancies"
    
    # Header row
    headers = ['Account Code', 'Vendor', 'Invoice #', 'Date', 'Amount', 'Issue Type', 'RVW Amount', 'Craftable Amount', 'Action']
    ws.append(headers)
    
    # Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Data rows
    for result in unmatched:
        # Determine vendor and issue type display
        vendor = result.vendor_rvw or result.vendor_craftable or '(Unknown)'

        # A synthetic "Total" row (gl_code == 'TOTAL') represents the whole
        # split invoice checked at the Total $ line rather than one GL line -
        # per BK's spec, this exports as an entire-invoice issue, not a
        # single-GL discrepancy.
        is_invoice_total = result.gl_code == 'TOTAL'

        if is_invoice_total:
            account_code = '(Entire Invoice)'
        else:
            account_code = result.gl_code
            if result.gl_description:
                account_code = f"{result.gl_code} - {result.gl_description}"

        if result.match_type == 'missing_rvw':
            issue_type = 'Entire Invoice Missing from RVW' if is_invoice_total else 'Missing from RVW'
            rvw_amt = ''
            craft_amt = f"${result.amount_craftable:.2f}"
            action_hint = 'accrue'
        elif result.match_type == 'missing_craftable':
            issue_type = 'Entire Invoice Missing from Craftable' if is_invoice_total else 'Missing from Craftable'
            rvw_amt = f"${result.amount_rvw:.2f}"
            craft_amt = ''
            action_hint = 'missing'
        elif result.fuzzy_match:
            issue_type = f'Possible reclass: ${result.difference:.2f}'
            rvw_amt = f"${result.amount_rvw:.2f}"
            craft_amt = f"${result.amount_craftable:.2f}"
            action_hint = 'reclass'
        else:  # amount_diff, same GL code both sides (or invoice total variance)
            issue_type = f'Entire Invoice Diff: ${result.difference:.2f}' if is_invoice_total else f'Diff: ${result.difference:.2f}'
            rvw_amt = f"${result.amount_rvw:.2f}"
            craft_amt = f"${result.amount_craftable:.2f}"
            action_hint = ''

        # Determine display amount
        display_amount = result.amount_rvw or result.amount_craftable

        row = [
            account_code,
            vendor,
            result.invoice_number,
            result.date,
            f"${display_amount:.2f}",
            issue_type,
            rvw_amt,
            craft_amt,
            action_hint  # Action column - pre-filled suggestion, editable by user
        ]
        
        ws.append(row)
    
    # Format data rows
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in row_cells:
            cell.border = thin_border
            if cell.column in [5, 7, 8]:  # Amount columns
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Auto-size columns
    column_widths = {
        'A': 24,  # Account Code (now includes description)
        'B': 25,  # Vendor
        'C': 15,  # Invoice #
        'D': 12,  # Date
        'E': 15,  # Amount
        'F': 24,  # Issue Type
        'G': 15,  # RVW Amount
        'H': 15,  # Craftable Amount
        'I': 20,  # Action
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to buffer
    wb.save(buffer)


def export_selected_results_to_excel(
    buffer,
    results: List[MatchResult],
    selected_keys: List[str],
    split_invoices: Optional[List[Dict]] = None
) -> None:
    """
    Export only selected reconciliation results to Excel.

    A split invoice's Total $ line uses the checkbox key "{invoice}_TOTAL"
    (see results.html). Checking it exports the discrepancy as a single
    entire-invoice row (RVW invoice total vs. Craftable invoice total)
    instead of any individual GL line, per BK's spec: check the Total to
    flag the whole invoice as missing/off, or check a specific GL line to
    flag it for reclassing.

    Args:
        buffer: BytesIO buffer to write Excel file to
        results: List of MatchResult objects
        selected_keys: List of selected result keys (format: "invoice_glcode"
            for GL lines, "invoice_TOTAL" for a split invoice's Total row)
        split_invoices: Split invoice dicts from Reconciler.get_split_invoices(),
            needed to resolve "invoice_TOTAL" keys
    """

    # Filter to selected GL-line results
    selected = [
        r for r in results
        if f"{r.invoice_number}_{r.gl_code}" in selected_keys
    ]

    # Resolve any selected "invoice_TOTAL" keys to synthetic invoice-level rows
    if split_invoices:
        splits_by_invoice = {s['invoice_number']: s for s in split_invoices}
        for key in selected_keys:
            if not key.endswith('_TOTAL'):
                continue
            invoice_number = key[:-len('_TOTAL')]
            split = splits_by_invoice.get(invoice_number)
            if split is None:
                continue
            selected.append(MatchResult(
                invoice_number=split['invoice_number'],
                gl_code='TOTAL',
                gl_description='',
                vendor_rvw=split['vendor'],
                vendor_craftable=split['vendor'],
                amount_rvw=split['total_rvw'],
                amount_craftable=split['total_craftable'],
                date=split['date'],
                match_type=split['total_match_type'],
                difference=split['total_difference'],
                variance_pct=0.0,
                is_split=True,
                split_group_id=split['invoice_number'],
                fuzzy_match=False,
                invoice_total_rvw=split['total_rvw'],
                invoice_total_craftable=split['total_craftable'],
            ))

    # Export filtered results
    export_results_to_excel(buffer, selected)
