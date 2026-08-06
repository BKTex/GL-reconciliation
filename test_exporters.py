"""
Unit tests for exporters.py module.

Covers exporting selected reconciliation results, including the split-invoice
Total $ checkbox, which must export as a single entire-invoice row rather than
any individual GL line.
"""

import io
import pytest
from openpyxl import load_workbook
from reconciler import MatchResult
from exporters import export_selected_results_to_excel


def make_result(**kwargs):
    defaults = dict(
        invoice_number="INV1",
        gl_code="40100",
        gl_description="Grocery",
        vendor_rvw="SYSCO",
        vendor_craftable="SYSCO",
        amount_rvw=100.0,
        amount_craftable=100.0,
        date="2026-01-01",
        match_type="matched",
        difference=0.0,
        variance_pct=0.0,
        is_split=False,
        split_group_id="INV1",
        fuzzy_match=False,
        invoice_total_rvw=100.0,
        invoice_total_craftable=100.0,
    )
    defaults.update(kwargs)
    return MatchResult(**defaults)


def read_rows(buffer):
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active
    return [tuple(cell.value for cell in row) for row in ws.iter_rows(min_row=2)]


class TestSelectedExport:

    def test_selected_gl_line_is_exported(self):
        results = [
            make_result(invoice_number="INV1", gl_code="40100", match_type="amount_diff",
                        amount_rvw=100.0, amount_craftable=90.0, difference=10.0),
        ]
        buffer = io.BytesIO()
        export_selected_results_to_excel(buffer, results, ["INV1_40100"])
        rows = read_rows(buffer)
        assert len(rows) == 1
        assert rows[0][0] == "40100 - Grocery"

    def test_unselected_line_is_excluded(self):
        results = [
            make_result(invoice_number="INV1", gl_code="40100", match_type="amount_diff",
                        difference=10.0),
            make_result(invoice_number="INV2", gl_code="40200", match_type="missing_rvw",
                        difference=-50.0),
        ]
        buffer = io.BytesIO()
        export_selected_results_to_excel(buffer, results, ["INV1_40100"])
        rows = read_rows(buffer)
        assert len(rows) == 1
        assert rows[0][2] == "INV1"


class TestSplitInvoiceTotalExport:
    """Checking a split invoice's Total $ box exports the discrepancy as a
    single entire-invoice row, not any of its individual GL lines."""

    def _split(self, **kwargs):
        defaults = dict(
            invoice_number="SPLIT1",
            vendor="SYSCO",
            date="2026-03-01",
            total_rvw=1000.0,
            total_craftable=950.0,
            total_difference=50.0,
            total_match_type="amount_diff",
            lines=[
                make_result(invoice_number="SPLIT1", gl_code="40100", is_split=True,
                             match_type="matched"),
                make_result(invoice_number="SPLIT1", gl_code="40200", is_split=True,
                             match_type="matched"),
            ],
        )
        defaults.update(kwargs)
        return defaults

    def test_total_checkbox_exports_single_entire_invoice_row(self):
        split_invoices = [self._split()]
        results = split_invoices[0]["lines"]

        buffer = io.BytesIO()
        export_selected_results_to_excel(
            buffer, results, ["SPLIT1_TOTAL"], split_invoices=split_invoices
        )
        rows = read_rows(buffer)

        assert len(rows) == 1
        account_code, vendor, invoice_number, date, amount, issue_type, rvw_amt, craft_amt, action = rows[0]
        assert account_code == "(Entire Invoice)"
        assert invoice_number == "SPLIT1"
        assert "Entire Invoice Diff" in issue_type

    def test_total_checkbox_missing_craftable_labels_entire_invoice(self):
        split_invoices = [self._split(
            total_craftable=0.0,
            total_difference=1000.0,
            total_match_type="missing_craftable",
        )]
        results = split_invoices[0]["lines"]

        buffer = io.BytesIO()
        export_selected_results_to_excel(
            buffer, results, ["SPLIT1_TOTAL"], split_invoices=split_invoices
        )
        rows = read_rows(buffer)

        assert len(rows) == 1
        issue_type = rows[0][5]
        assert issue_type == "Entire Invoice Missing from Craftable"

    def test_gl_line_checkbox_still_exports_individual_line_not_total(self):
        split_invoices = [self._split()]
        results = split_invoices[0]["lines"]
        # Force one line to be a discrepancy so it's a meaningful export target
        results[0] = make_result(invoice_number="SPLIT1", gl_code="40100", is_split=True,
                                  match_type="amount_diff", difference=25.0)

        buffer = io.BytesIO()
        export_selected_results_to_excel(
            buffer, results, ["SPLIT1_40100"], split_invoices=split_invoices
        )
        rows = read_rows(buffer)

        assert len(rows) == 1
        assert rows[0][0] != "(Entire Invoice)"
        assert rows[0][0].startswith("40100")

    def test_total_and_gl_line_both_checked_exports_both(self):
        split_invoices = [self._split()]
        results = split_invoices[0]["lines"]
        results[0] = make_result(invoice_number="SPLIT1", gl_code="40100", is_split=True,
                                  match_type="amount_diff", difference=25.0)

        buffer = io.BytesIO()
        export_selected_results_to_excel(
            buffer, results, ["SPLIT1_TOTAL", "SPLIT1_40100"], split_invoices=split_invoices
        )
        rows = read_rows(buffer)

        assert len(rows) == 2
        account_codes = {row[0] for row in rows}
        assert "(Entire Invoice)" in account_codes
